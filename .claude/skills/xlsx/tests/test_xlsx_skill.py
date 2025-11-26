"""
Comprehensive test suite for XLSX skill integration.

Tests are organized into 4 levels:
- Level 1: Skill Load Test (verify SKILL.md exists and is valid)
- Level 2: Dependency Test (check if dependencies are installed)
- Level 3: Basic Functionality Test (verify core operations work)
- Level 4: Integration Test (test skill in realistic scenarios)

Tests skip gracefully if dependencies are not installed.
"""

import json
import subprocess
import sys
import tempfile
from pathlib import Path

import pytest


# Test helpers
def check_dependency(package_name: str) -> bool:
    """Check if a Python package is installed."""
    try:
        __import__(package_name)
        return True
    except ImportError:
        return False


def check_command(command: str) -> bool:
    """Check if a system command is available."""
    try:
        subprocess.run([command, "--version"], capture_output=True, timeout=5)
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


# Dependency checks
HAS_PANDAS = check_dependency("pandas")
HAS_OPENPYXL = check_dependency("openpyxl")
HAS_LIBREOFFICE = check_command("soffice")
SKILL_DIR = Path(__file__).parent.parent


# ============================================================================
# LEVEL 1: SKILL LOAD TESTS
# ============================================================================


class TestLevel1SkillLoad:
    """Level 1: Verify SKILL.md exists and is valid."""

    def test_skill_file_exists(self):
        """Verify SKILL.md exists in the skill directory."""
        skill_file = SKILL_DIR / "SKILL.md"
        assert skill_file.exists(), f"SKILL.md not found at {skill_file}"

    def test_skill_file_readable(self):
        """Verify SKILL.md can be read."""
        skill_file = SKILL_DIR / "SKILL.md"
        content = skill_file.read_text()
        assert len(content) > 0, "SKILL.md is empty"

    def test_skill_yaml_frontmatter(self):
        """Verify SKILL.md has valid YAML frontmatter."""
        skill_file = SKILL_DIR / "SKILL.md"
        content = skill_file.read_text()

        assert content.startswith("---"), (
            "SKILL.md missing YAML frontmatter start delimiter"
        )

        # Find the closing ---
        lines = content.split("\n")
        yaml_end = None
        for idx, line in enumerate(lines[1:], start=1):
            if line.strip() == "---":
                yaml_end = idx
                break

        assert yaml_end is not None, "SKILL.md missing YAML frontmatter end delimiter"

    def test_skill_yaml_content(self):
        """Verify YAML frontmatter contains required fields."""
        import yaml

        skill_file = SKILL_DIR / "SKILL.md"
        content = skill_file.read_text()

        # Extract YAML
        parts = content.split("---")
        assert len(parts) >= 3, "Invalid YAML structure"

        yaml_content = parts[1]
        metadata = yaml.safe_load(yaml_content)

        # Verify required fields
        assert "name" in metadata, "YAML missing 'name' field"
        assert metadata["name"] == "xlsx", (
            f"Expected name 'xlsx', got '{metadata['name']}'"
        )
        assert "description" in metadata, "YAML missing 'description' field"
        assert "license" in metadata, "YAML missing 'license' field"

    def test_readme_exists(self):
        """Verify README.md exists with integration notes."""
        readme = SKILL_DIR / "README.md"
        assert readme.exists(), "README.md not found"

        content = readme.read_text()
        assert "amplihack" in content.lower(), "README missing amplihack context"

    def test_dependencies_doc_exists(self):
        """Verify DEPENDENCIES.md exists."""
        deps_file = SKILL_DIR / "DEPENDENCIES.md"
        assert deps_file.exists(), "DEPENDENCIES.md not found"

        content = deps_file.read_text()
        assert "pandas" in content, "DEPENDENCIES.md missing pandas"
        assert "openpyxl" in content, "DEPENDENCIES.md missing openpyxl"
        assert "LibreOffice" in content, "DEPENDENCIES.md missing LibreOffice"

    def test_recalc_script_exists(self):
        """Verify recalc.py script exists and is executable."""
        recalc_script = SKILL_DIR / "scripts" / "recalc.py"
        assert recalc_script.exists(), "recalc.py script not found"

        # Check if executable (Unix-like systems)
        if sys.platform != "win32":
            import os

            assert os.access(recalc_script, os.X_OK), "recalc.py is not executable"

    def test_examples_exist(self):
        """Verify examples directory and example_usage.md exist."""
        examples_dir = SKILL_DIR / "examples"
        assert examples_dir.exists(), "examples directory not found"

        example_file = examples_dir / "example_usage.md"
        assert example_file.exists(), "example_usage.md not found"

        content = example_file.read_text()
        assert "Example" in content, "example_usage.md appears to be empty or invalid"


# ============================================================================
# LEVEL 2: DEPENDENCY TESTS
# ============================================================================


class TestLevel2Dependencies:
    """Level 2: Verify dependencies are installed."""

    def test_pandas_installed(self):
        """Check if pandas is installed."""
        assert HAS_PANDAS, "pandas is not installed. Install with: pip install pandas"

    def test_openpyxl_installed(self):
        """Check if openpyxl is installed."""
        assert HAS_OPENPYXL, (
            "openpyxl is not installed. Install with: pip install openpyxl"
        )

    def test_libreoffice_available(self):
        """Check if LibreOffice is available."""
        if not HAS_LIBREOFFICE:
            pytest.skip(
                "LibreOffice not installed. See DEPENDENCIES.md for installation instructions."
            )

    def test_python_version(self):
        """Verify Python version is 3.8+."""
        assert sys.version_info >= (3, 8), (
            f"Python 3.8+ required, got {sys.version_info}"
        )

    @pytest.mark.skipif(not HAS_PANDAS, reason="pandas not installed")
    def test_pandas_version(self):
        """Verify pandas version is adequate."""
        import pandas as pd

        version = tuple(map(int, pd.__version__.split(".")[:2]))
        assert version >= (1, 5), f"pandas 1.5.0+ required, got {pd.__version__}"

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_openpyxl_version(self):
        """Verify openpyxl version is adequate."""
        import openpyxl

        version = tuple(map(int, openpyxl.__version__.split(".")[:2]))
        assert version >= (3, 0), (
            f"openpyxl 3.0.0+ required, got {openpyxl.__version__}"
        )


# ============================================================================
# LEVEL 3: BASIC FUNCTIONALITY TESTS
# ============================================================================


@pytest.mark.skipif(
    not (HAS_PANDAS and HAS_OPENPYXL), reason="pandas and openpyxl required"
)
class TestLevel3BasicFunctionality:
    """Level 3: Test basic XLSX operations."""

    def test_create_simple_workbook(self):
        """Test creating a basic Excel workbook with openpyxl."""
        from openpyxl import Workbook

        wb = Workbook()
        sheet = wb.active
        sheet["A1"] = "Test"
        sheet["B1"] = 123

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb.save(tmp_path)
            assert tmp_path.exists(), "Failed to create Excel file"
            assert tmp_path.stat().st_size > 0, "Created Excel file is empty"
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_create_workbook_with_formula(self):
        """Test creating a workbook with formulas."""
        from openpyxl import Workbook

        wb = Workbook()
        sheet = wb.active
        sheet["A1"] = 100
        sheet["A2"] = 200
        sheet["A3"] = "=SUM(A1:A2)"  # Formula

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb.save(tmp_path)
            assert tmp_path.exists(), "Failed to create Excel file with formulas"
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_load_and_modify_workbook(self):
        """Test loading and modifying an existing workbook."""
        from openpyxl import Workbook, load_workbook

        # Create initial workbook
        wb = Workbook()
        sheet = wb.active
        sheet["A1"] = "Original"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb.save(tmp_path)

            # Load and modify
            wb2 = load_workbook(tmp_path)
            sheet2 = wb2.active
            assert sheet2["A1"].value == "Original", "Failed to read original value"

            sheet2["A1"] = "Modified"
            wb2.save(tmp_path)

            # Verify modification
            wb3 = load_workbook(tmp_path)
            sheet3 = wb3.active
            assert sheet3["A1"].value == "Modified", "Failed to modify value"
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_pandas_read_write(self):
        """Test reading and writing Excel files with pandas."""
        import pandas as pd

        # Create test data
        df = pd.DataFrame(
            {
                "Name": ["Alice", "Bob", "Charlie"],
                "Age": [25, 30, 35],
                "Salary": [50000, 60000, 70000],
            }
        )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Write
            df.to_excel(tmp_path, index=False)
            assert tmp_path.exists(), "Failed to write Excel file with pandas"

            # Read back
            df2 = pd.read_excel(tmp_path)
            assert len(df2) == 3, "Failed to read correct number of rows"
            assert list(df2.columns) == ["Name", "Age", "Salary"], (
                "Failed to read correct columns"
            )
            assert df2["Name"].tolist() == ["Alice", "Bob", "Charlie"], (
                "Failed to read correct data"
            )
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_formula_preservation(self):
        """Test that formulas are preserved when loading and saving."""
        from openpyxl import Workbook, load_workbook

        wb = Workbook()
        sheet = wb.active
        sheet["A1"] = 10
        sheet["A2"] = 20
        sheet["A3"] = "=A1+A2"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb.save(tmp_path)

            # Load without data_only to preserve formulas
            wb2 = load_workbook(tmp_path, data_only=False)
            sheet2 = wb2.active

            # Verify formula is preserved
            assert sheet2["A3"].value == "=A1+A2", "Formula was not preserved"
        finally:
            tmp_path.unlink(missing_ok=True)

    @pytest.mark.skipif(not HAS_LIBREOFFICE, reason="LibreOffice not installed")
    def test_recalc_script_basic(self):
        """Test recalc.py script with a simple workbook."""
        from openpyxl import Workbook

        wb = Workbook()
        sheet = wb.active
        sheet["A1"] = 10
        sheet["A2"] = 20
        sheet["A3"] = "=A1+A2"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb.save(tmp_path)

            # Run recalc.py
            recalc_script = SKILL_DIR / "scripts" / "recalc.py"
            result = subprocess.run(
                [sys.executable, str(recalc_script), str(tmp_path)],
                capture_output=True,
                text=True,
                timeout=30,
            )

            assert result.returncode == 0, f"recalc.py failed: {result.stderr}"

            # Parse JSON output
            output = json.loads(result.stdout)
            assert "status" in output, "recalc.py output missing 'status' field"
            assert output["status"] == "success", f"recalc.py reported errors: {output}"
            assert output["total_errors"] == 0, (
                f"Expected 0 errors, got {output['total_errors']}"
            )
        finally:
            tmp_path.unlink(missing_ok=True)


# ============================================================================
# LEVEL 4: INTEGRATION TESTS
# ============================================================================


@pytest.mark.skipif(
    not (HAS_PANDAS and HAS_OPENPYXL), reason="pandas and openpyxl required"
)
class TestLevel4Integration:
    """Level 4: Test realistic usage scenarios."""

    def test_financial_model_creation(self):
        """Test creating a simple financial model with formulas."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Financial Model"

        # Headers
        sheet["A1"] = "Item"
        sheet["B1"] = "Amount"
        sheet["A1"].font = Font(bold=True)
        sheet["B1"].font = Font(bold=True)

        # Data with formulas
        sheet["A2"] = "Revenue"
        sheet["B2"] = 100000
        sheet["B2"].font = Font(color="0000FF")  # Blue for input

        sheet["A3"] = "Cost of Sales"
        sheet["B3"] = "=-B2*0.4"  # Formula
        sheet["B3"].font = Font(color="000000")  # Black for formula

        sheet["A4"] = "Gross Profit"
        sheet["B4"] = "=B2+B3"  # Formula (B3 is negative)
        sheet["B4"].font = Font(color="000000")

        sheet["A5"] = "Operating Expenses"
        sheet["B5"] = "=-B2*0.25"
        sheet["B5"].font = Font(color="000000")

        sheet["A6"] = "EBITDA"
        sheet["B6"] = "=B4+B5"
        sheet["B6"].font = Font(bold=True, color="000000")

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb.save(tmp_path)
            assert tmp_path.exists(), "Failed to create financial model"

            # Verify formulas are present
            from openpyxl import load_workbook

            wb2 = load_workbook(tmp_path, data_only=False)
            sheet2 = wb2.active

            assert sheet2["B3"].value == "=-B2*0.4", "Cost of Sales formula missing"
            assert sheet2["B4"].value == "=B2+B3", "Gross Profit formula missing"
            assert sheet2["B6"].value == "=B4+B5", "EBITDA formula missing"
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_multi_sheet_workbook(self):
        """Test creating a workbook with multiple sheets and cross-sheet references."""
        from openpyxl import Workbook

        wb = Workbook()

        # Sheet 1: Data
        data_sheet = wb.active
        data_sheet.title = "Data"
        data_sheet["A1"] = "Value"
        data_sheet["B1"] = 1000

        # Sheet 2: Summary with reference to Sheet 1
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet["A1"] = "Total"
        summary_sheet["B1"] = "=Data!B1*2"  # Cross-sheet reference

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb.save(tmp_path)

            # Verify multiple sheets
            from openpyxl import load_workbook

            wb2 = load_workbook(tmp_path)
            assert len(wb2.sheetnames) == 2, "Expected 2 sheets"
            assert "Data" in wb2.sheetnames, "Data sheet missing"
            assert "Summary" in wb2.sheetnames, "Summary sheet missing"

            # Verify cross-sheet reference
            summary = wb2["Summary"]
            assert summary["B1"].value == "=Data!B1*2", "Cross-sheet reference missing"
        finally:
            tmp_path.unlink(missing_ok=True)

    @pytest.mark.skipif(not HAS_LIBREOFFICE, reason="LibreOffice not installed")
    def test_error_detection(self):
        """Test that recalc.py detects formula errors."""
        from openpyxl import Workbook

        wb = Workbook()
        sheet = wb.active

        # Create a formula with intentional errors
        sheet["A1"] = 10
        sheet["A2"] = 0
        sheet["A3"] = "=A1/A2"  # Division by zero
        sheet["A4"] = "=A1+B99"  # Reference to empty cell (not an error)
        sheet["A5"] = "=INVALID(A1)"  # Invalid function name

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb.save(tmp_path)

            # Run recalc.py
            recalc_script = SKILL_DIR / "scripts" / "recalc.py"
            result = subprocess.run(
                [sys.executable, str(recalc_script), str(tmp_path)],
                capture_output=True,
                text=True,
                timeout=30,
            )

            # Parse output
            output = json.loads(result.stdout)

            # Should detect errors
            assert output["status"] == "errors_found", "Expected errors to be found"
            assert output["total_errors"] > 0, "Expected at least one error"

            # Check for specific error types
            if "error_summary" in output:
                # We expect #DIV/0! and #NAME? errors
                error_types = output["error_summary"].keys()
                # At least one of these should be present
                assert len(error_types) > 0, "Expected error types in summary"
        finally:
            tmp_path.unlink(missing_ok=True)

    def test_data_analysis_workflow(self):
        """Test a complete data analysis workflow with pandas."""
        import pandas as pd

        # Create sample data
        df = pd.DataFrame(
            {
                "Product": ["A", "B", "C", "A", "B", "C"],
                "Region": ["East", "East", "East", "West", "West", "West"],
                "Sales": [1000, 1500, 1200, 1100, 1600, 1300],
            }
        )

        # Analyze
        summary = df.groupby("Product")["Sales"].sum().reset_index()
        summary.columns = ["Product", "Total Sales"]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            # Export to Excel with multiple sheets
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Raw Data", index=False)
                summary.to_excel(writer, sheet_name="Summary", index=False)

            assert tmp_path.exists(), "Failed to create Excel file"

            # Read back and verify
            df_read = pd.read_excel(tmp_path, sheet_name="Raw Data")
            summary_read = pd.read_excel(tmp_path, sheet_name="Summary")

            assert len(df_read) == 6, "Expected 6 rows in raw data"
            assert len(summary_read) == 3, "Expected 3 products in summary"
            assert summary_read["Total Sales"].sum() == 7700, "Incorrect total sales"
        finally:
            tmp_path.unlink(missing_ok=True)

    @pytest.mark.skipif(not HAS_LIBREOFFICE, reason="LibreOffice not installed")
    def test_complete_workflow(self):
        """Test complete workflow: create, add formulas, recalculate, verify."""
        from openpyxl import Workbook, load_workbook

        # Step 1: Create workbook with formulas
        wb = Workbook()
        sheet = wb.active

        sheet["A1"] = "Q1"
        sheet["A2"] = "Q2"
        sheet["A3"] = "Q3"
        sheet["A4"] = "Q4"
        sheet["A5"] = "Total"

        sheet["B1"] = 1000
        sheet["B2"] = 1200
        sheet["B3"] = 1100
        sheet["B4"] = 1300
        sheet["B5"] = "=SUM(B1:B4)"

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            wb.save(tmp_path)

            # Step 2: Recalculate formulas
            recalc_script = SKILL_DIR / "scripts" / "recalc.py"
            result = subprocess.run(
                [sys.executable, str(recalc_script), str(tmp_path)],
                capture_output=True,
                text=True,
                timeout=30,
            )

            assert result.returncode == 0, f"recalc.py failed: {result.stderr}"

            # Step 3: Verify zero errors
            output = json.loads(result.stdout)
            assert output["status"] == "success", f"Errors found: {output}"
            assert output["total_errors"] == 0, "Expected zero errors"
            assert output["total_formulas"] >= 1, "Expected at least one formula"

            # Step 4: Load and verify calculated values
            wb2 = load_workbook(tmp_path, data_only=True)
            sheet2 = wb2.active

            # After recalculation, B5 should contain the calculated value
            total = sheet2["B5"].value
            # Note: total might be None if LibreOffice didn't calculate,
            # but recalc.py should have set it
            if total is not None:
                assert total == 4600, f"Expected total 4600, got {total}"
        finally:
            tmp_path.unlink(missing_ok=True)


# ============================================================================
# TEST SUMMARY
# ============================================================================


def test_summary(capsys):
    """Print test summary information."""
    print("\n" + "=" * 70)
    print("XLSX SKILL TEST SUMMARY")
    print("=" * 70)
    print(f"Skill Directory: {SKILL_DIR}")
    print("\nDependency Status:")
    print(f"  pandas: {'✓ Installed' if HAS_PANDAS else '✗ Not Installed'}")
    print(f"  openpyxl: {'✓ Installed' if HAS_OPENPYXL else '✗ Not Installed'}")
    print(f"  LibreOffice: {'✓ Available' if HAS_LIBREOFFICE else '✗ Not Available'}")

    if not (HAS_PANDAS and HAS_OPENPYXL):
        print("\n⚠ Some tests will be skipped due to missing dependencies.")
        print("  Install dependencies: pip install pandas openpyxl")

    if not HAS_LIBREOFFICE:
        print(
            "\n⚠ Formula recalculation tests will be skipped (LibreOffice not found)."
        )
        print("  See DEPENDENCIES.md for LibreOffice installation instructions.")

    if HAS_PANDAS and HAS_OPENPYXL and HAS_LIBREOFFICE:
        print("\n✓ All dependencies available - full test suite will run.")

    print("=" * 70 + "\n")
